

import pandas as pd
from openpyxl import load_workbook

class Result_Handler():

    def __init__(self):
        pass

    def save_to_file(self, path, sheetName, df, append = False, header = False):
        try:
            book = load_workbook(path)
            writer = pd.ExcelWriter(path, engine='openpyxl')
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            startrow = 0
            if append == True:
                if sheetName in writer.sheets:
                    startrow = writer.sheets[sheetName].max_row
            df.to_excel(writer,sheet_name=sheetName, startrow=startrow, index = False, header= False)
            writer.save()
        except FileNotFoundError:
            with pd.ExcelWriter(path) as writer: 
                df.to_excel(writer, sheet_name=sheetName, header = header, index = False)





















